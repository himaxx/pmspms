"""
PMS Excel → JSON Extraction & Cleaning Script
Reads FMS master sheet, deduplicates, normalizes, outputs cleaned_jobs.json
"""
import json, sys, io, os
from datetime import datetime, timedelta
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'PMS FORM RESPONSES.xlsx')

# FMS Column mapping (Row 4 = headers, data starts Row 5)
FMS_COL = {
    0: 'job_no', 1: 'date', 2: 'prog_by', 3: 'item', 4: 'size',
    5: 'qty', 6: 'reason', 7: 'special_instruction', 8: 'item_group',
    9: 's2_planned', 10: 's2_actual', 11: 's2_yes_no',
    12: 's2_inhouse', 13: 's2_instructions', 14: 's2_delay',
    15: 's3_planned', 16: 's3_actual', 17: 's3_dukan_cutting',
    18: 's3_size_details', 19: 's3_cutting_person', 20: 's3_delay',
    21: 's2_outside_cutting_planned', 22: 's3_inhouse_planned',
    23: 's4_planned', 24: 's4_start_date', 25: 's4_thekedar',
    26: 's4_cutting_pcs', 27: 's4_cut_to_pack', 28: 's4_lead_time',
    29: 's4_vastra_job', 30: 's4_inhouse_cutting_ref',
    31: 's4_fixed_cutting', 32: 's4_open_cutting', 33: 's4_delay',
    34: 's5_unfinished_jama', 35: 's5_lead_time_hours',
    36: 's5_jama_planned', 37: 's5_jobslip_status', 38: 's5_status',
    39: 's5_balance', 40: 's5_jama_qty', 41: 's5_given_qty',
    42: 's5_press', 43: 's5_delay', 44: 's6_settle_qty',
    45: 's6_reason', 46: 's6_name',
}

# Item group normalization map
ITEM_GROUP_MAP = {
    'full bottom': 'Full Bottom',
    'skirts': 'Skirt', 'skirt': 'Skirt',
    'shorts': 'Shorts',
    'capri': 'Capri',
    'tops / tshirts': 'Tops', 'tops': 'Tops', 'tops/tshirts': 'Tops',
    'aline / frock / long tops': 'Alin/Frock/Long Tops',
    'alin/frock/long tops': 'Alin/Frock/Long Tops',
    'sets': 'Sets',
    'boys': 'Boys',
}

# Thekedar name normalization
THEKEDAR_MAP = {
    'mukati ji': 'Mukati Ji', 'sitaram ji': 'Sitaram Ji',
    'jitendra chore ji': 'Jitendre Chorey Ji',
    'jitendre chorey ji': 'Jitendre Chorey Ji',
    'anil panwar': 'Anil Pawar Ji', 'kadir': 'Abdul Kadir Ji',
    'kalu rajpoot': 'Kalu Rajpoot', 'kalu sig': 'Kalu Sig',
    'pappu': 'Pappu', 'ravi panwar': 'Ravi Panwar', 'xyz': 'Xyz',
}

# Infer item_group from prog_by (deterministic 1:1 mappings only)
PROGBY_TO_ITEMGROUP = {
    'Capri - Sanjay': 'Capri',
    'Shorts - Manohar': 'Shorts',
    'Skirts - Ankit': 'Skirt',
    'Tops Hosiery - Naresh': 'Tops',
}

# Date columns that should be ISO timestamps
DATE_COLS = {
    'date', 's2_planned', 's2_actual', 's3_planned', 's3_actual',
    's2_outside_cutting_planned', 's3_inhouse_planned',
    's4_planned', 's4_start_date', 's5_jama_planned', 's5_jobslip_status',
}

# Numeric columns
NUM_COLS = {
    'qty', 's3_dukan_cutting', 's4_cutting_pcs', 's4_lead_time',
    's4_inhouse_cutting_ref', 's4_fixed_cutting', 's4_open_cutting',
    's5_lead_time_hours', 's5_balance', 's5_jama_qty', 's5_given_qty',
    's6_settle_qty',
}


def safe_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None


def safe_int(v):
    if v is None: return None
    try:
        f = float(str(v).strip())
        return int(f)
    except (ValueError, TypeError):
        return None


def safe_delay_hours(v):
    """Convert duration strings like '1:05:02', '22 days, 4:21:26.529000' to numeric hours."""
    if v is None: return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip()
    if not s: return None
    
    # Try parsing as a plain number first
    try:
        return round(float(s), 2)
    except ValueError:
        pass
    
    # Parse "X days, H:M:S" or "H:M:S" format
    import re
    days = 0
    negative = False
    if s.startswith('-'):
        negative = True
        s = s[1:]
    
    day_match = re.match(r'(\d+)\s*days?,\s*(.*)', s)
    if day_match:
        days = int(day_match.group(1))
        s = day_match.group(2)
    
    # Parse H:M:S or H:M:S.microseconds
    time_match = re.match(r'(\d+):(\d+):(\d+)', s)
    if time_match:
        h = int(time_match.group(1))
        m = int(time_match.group(2))
        sec = int(time_match.group(3))
        total_hours = days * 24 + h + m / 60 + sec / 3600
        return round(-total_hours if negative else total_hours, 2)
    
    return None


def safe_iso(v):
    if v is None: return None
    if isinstance(v, datetime):
        return v.isoformat()
    s = str(v).strip()
    if not s: return None
    try:
        d = datetime.fromisoformat(s.replace('Z', '+00:00'))
        if d.year < 1990: return None
        return d.isoformat()
    except ValueError:
        try:
            d = datetime.strptime(s, '%Y-%m-%d %H:%M:%S.%f')
            return d.isoformat()
        except ValueError:
            try:
                d = datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
                return d.isoformat()
            except ValueError:
                return None


def normalize_job_no(v):
    if v is None: return None
    s = str(v).strip()
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s if s else None


def normalize_item_group(v):
    if not v: return None
    return ITEM_GROUP_MAP.get(str(v).strip().lower(), str(v).strip())


def normalize_thekedar(v):
    if not v: return None
    s = str(v).strip()
    return THEKEDAR_MAP.get(s.lower(), s)


def normalize_prog_by(v):
    """Normalize 'Tops Hosiery - Naresh' or 'Capri - Sanjay' to just the name."""
    if not v: return None
    s = str(v).strip()
    if ' - ' in s:
        parts = s.split(' - ')
        return parts[-1].strip()
    return s


def row_completeness(record):
    """Count non-null fields for deduplication ranking."""
    return sum(1 for v in record.values() if v is not None)


def extract_fms(wb):
    """Extract FMS master sheet, filtering empty placeholders."""
    ws = wb['FMS']
    all_rows = list(ws.iter_rows(min_row=5, values_only=True))  # Data starts row 5
    print(f"  FMS raw rows: {len(all_rows)}")

    # Parse all rows into records
    records_by_job = {}
    skipped = 0
    empty_placeholder = 0

    for row in all_rows:
        job_no = normalize_job_no(row[0] if len(row) > 0 else None)
        if not job_no:
            skipped += 1
            continue

        # Filter: skip rows without item name AND date (empty placeholders)
        item_val = row[3] if len(row) > 3 else None
        date_val = row[1] if len(row) > 1 else None
        has_item = item_val is not None and str(item_val).strip() != ''
        has_date = date_val is not None and str(date_val).strip() != ''
        if not has_item and not has_date:
            empty_placeholder += 1
            continue

        record = {}
        for col_idx, col_name in FMS_COL.items():
            val = row[col_idx] if col_idx < len(row) else None

            if col_name == 'job_no':
                record[col_name] = job_no
            elif col_name in DATE_COLS:
                record[col_name] = safe_iso(val)
            elif col_name in NUM_COLS:
                record[col_name] = safe_int(val)
            elif col_name == 'item_group':
                record[col_name] = normalize_item_group(val)
            elif col_name == 's4_thekedar':
                record[col_name] = normalize_thekedar(val)
            elif col_name == 'prog_by':
                record[col_name] = normalize_prog_by(val)
            elif col_name in ('s2_delay', 's3_delay', 's4_delay', 's5_delay'):
                record[col_name] = safe_delay_hours(val)
            else:
                record[col_name] = safe_str(val)

        # Infer item_group from prog_by if missing
        if not record.get('item_group') and record.get('prog_by'):
            inferred = PROGBY_TO_ITEMGROUP.get(record['prog_by'])
            if inferred:
                record['item_group'] = inferred

        # Deduplicate: keep the most complete record per job_no
        if job_no in records_by_job:
            existing = records_by_job[job_no]
            if row_completeness(record) > row_completeness(existing):
                for k, v in existing.items():
                    if record.get(k) is None and v is not None:
                        record[k] = v
                records_by_job[job_no] = record
            else:
                for k, v in record.items():
                    if existing.get(k) is None and v is not None:
                        existing[k] = v
        else:
            records_by_job[job_no] = record

    print(f"  FMS skipped (no job_no): {skipped}")
    print(f"  FMS empty placeholders filtered: {empty_placeholder}")
    print(f"  FMS real jobs extracted: {len(records_by_job)}")
    return records_by_job


def enrich_from_step_sheets(wb, jobs):
    """Cross-reference individual step sheets to fill gaps."""

    # Step 5: Finished Maal Jama — aggregate multiple jama entries
    ws5 = wb['Finished Maal Jama']
    jama_totals = {}
    press_status = {}
    for row in ws5.iter_rows(min_row=2, values_only=True):
        jno = normalize_job_no(row[0])
        if not jno: continue
        qty = safe_int(row[2])
        if qty:
            jama_totals[jno] = jama_totals.get(jno, 0) + qty
        press = safe_str(row[3]) if len(row) > 3 else None
        if press:
            press_status[jno] = 'Yes' if '✅' in press or 'Hogyi' in str(press) else 'No'

    updated_jama = 0
    for jno, total in jama_totals.items():
        if jno in jobs:
            if jobs[jno].get('s5_jama_qty') is None:
                jobs[jno]['s5_jama_qty'] = total
                updated_jama += 1
            if jno in press_status and jobs[jno].get('s5_press') is None:
                jobs[jno]['s5_press'] = press_status[jno]
    print(f"  Step 5 enrichment: {updated_jama} jama gaps filled")

    # Step 6: Settle — aggregate multiple settle entries
    ws6 = wb['Settle']
    settle_totals = {}
    settle_reasons = {}
    settle_names = {}
    for row in ws6.iter_rows(min_row=2, values_only=True):
        jno = normalize_job_no(row[1])
        if not jno: continue
        qty = safe_int(row[2])
        if qty: settle_totals[jno] = settle_totals.get(jno, 0) + qty
        reason = safe_str(row[3]) if len(row) > 3 else None
        name = safe_str(row[4]) if len(row) > 4 else None
        if reason: settle_reasons[jno] = reason
        if name: settle_names[jno] = name

    updated_settle = 0
    for jno, total in settle_totals.items():
        if jno in jobs:
            if jobs[jno].get('s6_settle_qty') is None:
                jobs[jno]['s6_settle_qty'] = total
                updated_settle += 1
            if jno in settle_reasons and jobs[jno].get('s6_reason') is None:
                jobs[jno]['s6_reason'] = settle_reasons[jno]
            if jno in settle_names and jobs[jno].get('s6_name') is None:
                jobs[jno]['s6_name'] = settle_names[jno]
    print(f"  Step 6 enrichment: {updated_settle} settle gaps filled")

    # Step 2: Production Approval — fill s2_approver
    ws2 = wb['Production Approval']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        jno = normalize_job_no(row[1])
        if not jno or jno not in jobs: continue
        if jobs[jno].get('s2_approver') is None:
            instructions = safe_str(row[4]) if len(row) > 4 else None
            if instructions and jobs[jno].get('s2_instructions') is None:
                jobs[jno]['s2_instructions'] = instructions

    # Step 4: Naame — fill thekedar gaps
    ws4 = wb['Naame']
    for row in ws4.iter_rows(min_row=2, values_only=True):
        jno = normalize_job_no(row[0])
        if not jno or jno not in jobs: continue
        if jobs[jno].get('s4_thekedar') is None:
            thekedar = normalize_thekedar(safe_str(row[2]))
            if thekedar: jobs[jno]['s4_thekedar'] = thekedar
        if jobs[jno].get('s4_cut_to_pack') is None:
            ctp = safe_str(row[3])
            if ctp: jobs[jno]['s4_cut_to_pack'] = ctp
        if jobs[jno].get('s4_lead_time') is None:
            lt = safe_int(row[4])
            if lt: jobs[jno]['s4_lead_time'] = lt
        if jobs[jno].get('s4_cutting_pcs') is None:
            cp = safe_int(row[5])
            if cp: jobs[jno]['s4_cutting_pcs'] = cp

    return jobs


def generate_stats(jobs):
    """Generate validation statistics."""
    stats = {
        'total_jobs': len(jobs),
        'by_item_group': {},
        'by_status': {'Complete': 0, 'In Progress': 0, 'Rejected': 0},
        'step_completion': {f'step_{i}': 0 for i in range(1, 7)},
        'date_range': {'earliest': None, 'latest': None},
    }

    for j in jobs.values():
        # Item groups
        ig = j.get('item_group', 'Unknown') or 'Unknown'
        stats['by_item_group'][ig] = stats['by_item_group'].get(ig, 0) + 1

        # Status
        if j.get('s5_status') == 'Complete':
            stats['by_status']['Complete'] += 1
        elif j.get('s2_yes_no') == 'No':
            stats['by_status']['Rejected'] += 1
        else:
            stats['by_status']['In Progress'] += 1

        # Step completion
        stats['step_completion']['step_1'] += 1  # All have step 1
        if j.get('s2_actual'): stats['step_completion']['step_2'] += 1
        if j.get('s3_actual'): stats['step_completion']['step_3'] += 1
        if j.get('s4_start_date'): stats['step_completion']['step_4'] += 1
        if j.get('s5_jama_qty'): stats['step_completion']['step_5'] += 1
        if j.get('s6_settle_qty') is not None: stats['step_completion']['step_6'] += 1

        # Date range
        d = j.get('date')
        if d:
            if not stats['date_range']['earliest'] or d < stats['date_range']['earliest']:
                stats['date_range']['earliest'] = d
            if not stats['date_range']['latest'] or d > stats['date_range']['latest']:
                stats['date_range']['latest'] = d

    return stats


def main():
    print("=" * 60)
    print("PMS EXCEL DATA EXTRACTION")
    print("=" * 60)

    print(f"\n📂 Loading: {os.path.abspath(EXCEL_PATH)}")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Phase 1: Extract & deduplicate FMS
    print("\n── Phase 1: FMS Extraction ──")
    jobs = extract_fms(wb)

    # Phase 2: Enrich from step sheets
    print("\n── Phase 2: Cross-Reference Enrichment ──")
    jobs = enrich_from_step_sheets(wb, jobs)

    # Add import metadata
    batch_id = datetime.now().strftime('excel_%Y%m%d_%H%M%S')
    for j in jobs.values():
        j['imported_from_excel'] = True
        j['import_batch_id'] = batch_id
        # Remove cols that don't exist in DB schema yet
        # Keep only known DB columns
        for key in ['s2_outside_cutting_planned', 's3_inhouse_planned',
                     's4_inhouse_cutting_ref', 's4_fixed_cutting',
                     's4_open_cutting', 's5_unfinished_jama']:
            j.pop(key, None)

    # Phase 3: Generate stats
    print("\n── Phase 3: Validation Stats ──")
    stats = generate_stats(jobs)
    print(f"  Total unique jobs: {stats['total_jobs']}")
    print(f"  By Item Group: {json.dumps(stats['by_item_group'], indent=4)}")
    print(f"  By Status: {json.dumps(stats['by_status'], indent=4)}")
    print(f"  Step Completion: {json.dumps(stats['step_completion'], indent=4)}")
    print(f"  Date Range: {stats['date_range']['earliest']} → {stats['date_range']['latest']}")

    # Phase 4: Output
    out_path = os.path.join(os.path.dirname(__file__), 'cleaned_jobs.json')
    jobs_list = sorted(jobs.values(), key=lambda x: int(x.get('job_no', '0') or '0'))
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(jobs_list, f, indent=2, ensure_ascii=False, default=str)

    stats_path = os.path.join(os.path.dirname(__file__), 'import_stats.json')
    with open(stats_path, 'w', encoding='utf-8') as f:
        json.dump(stats, f, indent=2, ensure_ascii=False)

    print(f"\n✅ Output: {os.path.abspath(out_path)} ({len(jobs_list)} records)")
    print(f"✅ Stats:  {os.path.abspath(stats_path)}")
    print("=" * 60)


if __name__ == '__main__':
    main()
